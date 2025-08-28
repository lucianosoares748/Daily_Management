import streamlit as st
from datetime import datetime
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

st.set_page_config(page_title="Daily Management", layout="centered")

st.title("📊 Daily Management - Cálculo de Indicadores")

# Seleção do técnico
tecnicos = ["Gustavo Galle", "Jonathan Giuli", "Diogo", "Dark"]
tecnico = st.selectbox("Selecione o Técnico", tecnicos)
data = st.date_input("Data", datetime.today())

st.subheader("⏱️ Informe os tempos em horímetro das máquinas")

maquinas = ["ES08", "ES09", "ES10", "ES11"]

TT_total = 0.0  # Tempo Trabalhado total
TM_total = 0.0  # Tempo de Manutenção total

col1, col2 = st.columns(2)

for maq in maquinas:
    with col1:
        trabalho = st.number_input(f"Tempo de TRABALHO ({maq})", min_value=0.0, step=0.1, format="%.1f")
    with col2:
        manutencao = st.number_input(f"Tempo de MANUTENÇÃO ({maq})", min_value=0.0, step=0.1, format="%.1f")

    TT_total += trabalho
    TM_total += manutencao

TP_por_maquina = st.number_input("TP (Tempo programado por máquina em horas)", value=8.8, step=0.1)
TP_total = TP_por_maquina * len(maquinas)

if st.button("Calcular"):
    # Cálculos
    TD_total = TP_total - TM_total
    DM = ((TP_total - TM_total) / TP_total) * 100 if TP_total > 0 else 0
    EO = (TT_total / TD_total) * 100 if TD_total > 0 else 0

    # Resumo formatado
    resumo = (
        f"Dia: {data.strftime('%d/%m/%Y')}\n"
        f"Técnico: {tecnico}\n"
        f"Horas Trabalhadas: {TT_total:.2f}\n"
        f"EO: {EO:.2f}%\n"
        f"Horas de Manutenção: {TM_total:.2f}\n"
        f"DM: {DM:.2f}%"
    )


    st.success("✅ Cálculo realizado com sucesso!")
    st.markdown("### 📋 Resumo do Dia")
    st.text_area("Resultado", resumo.strip(), height=160)

    # Botão copiar (JS)
    copy_code = f"""
    <script>
    function copyToClipboard(text) {{
        navigator.clipboard.writeText(text);
        alert("Resumo copiado para área de transferência!");
    }}
    </script>
    <button onclick="copyToClipboard(`{resumo.strip()}`)">📋 Copiar Resumo</button>
    """
    st.components.v1.html(copy_code, height=50)

    # --- Exportar para PDF ---
    def gerar_pdf(texto):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer)
        styles = getSampleStyleSheet()
        story = [Paragraph("Relatório Diário - Daily Management", styles['Title']), Spacer(1, 20)]
        for linha in texto.strip().split("\n"):
            story.append(Paragraph(linha.strip(), styles['Normal']))
            story.append(Spacer(1, 10))
        doc.build(story)
        buffer.seek(0)
        return buffer

    pdf_file = gerar_pdf(resumo)

    st.download_button(
        label="📄 Baixar PDF",
        data=pdf_file,
        file_name=f"relatorio_{data.strftime('%d-%m-%Y')}.pdf",
        mime="application/pdf"
    )

    # --- Exportar para Excel ---
    def gerar_excel(texto):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"

        for i, linha in enumerate(texto.strip().split("\n"), start=1):
            if ":" in linha:
                chave, valor = linha.split(":", 1)
                ws[f"A{i}"] = chave.strip()
                ws[f"B{i}"] = valor.strip()
            else:
                ws[f"A{i}"] = linha.strip()

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    excel_file = gerar_excel(resumo)

    st.download_button(
        label="📊 Baixar Excel",
        data=excel_file,
        file_name=f"relatorio_{data.strftime('%d-%m-%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
